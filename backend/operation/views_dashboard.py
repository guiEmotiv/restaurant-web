from rest_framework import viewsets, status
from rest_framework.response import Response
from rest_framework.decorators import action
from django.utils import timezone
from django.db.models import Sum, Count, Q, F, Avg
from django.http import JsonResponse
from datetime import datetime, date
from decimal import Decimal
from .models import Order, OrderItem, Payment

class DashboardViewSet(viewsets.ViewSet):
    """
    ViewSet para el dashboard consolidado - versión simplificada y funcional
    """
    permission_classes = []  # Acceso completo para usuarios autenticados
    
    @action(detail=False, methods=['get'])
    def report(self, request):
        """
        Endpoint consolidado para el dashboard con todos los datos finales
        Solo pedidos PAID, sin métricas en tiempo real
        """
        try:
            print(f"🔍 Dashboard report started - Request: {request.method} {request.path}")
            
            # Obtener fecha del parámetro o usar hoy (zona horaria Lima)
            date_param = request.query_params.get('date')
            if date_param:
                try:
                    selected_date = datetime.strptime(date_param, '%Y-%m-%d').date()
                    print(f"✅ Parsed date from parameter: {selected_date}")
                except ValueError:
                    selected_date = timezone.now().date()
                    print(f"⚠️ Invalid date format, using today: {selected_date}")
            else:
                selected_date = timezone.now().date()
                print(f"✅ Using current date: {selected_date}")
            
            print(f"🔍 Filtering orders for date: {selected_date}")
            
            # Filtrar órdenes PAID por fecha de paid_at
            print("🔍 Starting order query...")
            paid_orders = Order.objects.filter(
                status='PAID',
                paid_at__date=selected_date
            ).select_related(
                'table__zone'
            ).prefetch_related(
                'orderitem_set__recipe__group',
                'payments'
            ).order_by('paid_at')
            
            orders_count = paid_orders.count()
            print(f"✅ Found {orders_count} paid orders")
            
            # Métricas básicas del día
            print("🔍 Calculating basic metrics...")
            total_orders = orders_count
            total_revenue = paid_orders.aggregate(total=Sum('total_amount'))['total'] or Decimal('0')
            average_ticket = total_revenue / total_orders if total_orders > 0 else Decimal('0')
            
            print(f"✅ Metrics: {total_orders} orders, {total_revenue} revenue, {average_ticket} avg ticket")
            
            # Inicializar estadísticas
            category_stats = {}
            dish_stats = {}
            waiter_revenue = {}
            zone_revenue = {}
            table_revenue = {}
            payment_method_totals = {}
            service_times = []
            
            # Procesar cada orden pagada
            print(f"🔍 Processing {total_orders} orders...")
            
            for order in paid_orders:
                try:
                    # Calcular tiempo de servicio
                    if order.created_at and order.paid_at:
                        service_time_minutes = int((order.paid_at - order.created_at).total_seconds() / 60)
                        service_times.append(service_time_minutes)
                    
                    # Stats por mesero
                    waiter_key = order.waiter or 'Sin asignar'
                    if waiter_key not in waiter_revenue:
                        waiter_revenue[waiter_key] = {'orders': 0, 'revenue': Decimal('0')}
                    waiter_revenue[waiter_key]['orders'] += 1
                    waiter_revenue[waiter_key]['revenue'] += order.total_amount
                    
                    # Stats por zona
                    zone_key = order.table.zone.name if order.table and order.table.zone else 'Sin zona'
                    if zone_key not in zone_revenue:
                        zone_revenue[zone_key] = {'orders': 0, 'revenue': Decimal('0'), 'tables': set()}
                    zone_revenue[zone_key]['orders'] += 1
                    zone_revenue[zone_key]['revenue'] += order.total_amount
                    if order.table:
                        zone_revenue[zone_key]['tables'].add(order.table.number)
                    
                    # Stats por mesa
                    if order.table:
                        table_key = f"Mesa {order.table.number}"
                        table_revenue[table_key] = table_revenue.get(table_key, Decimal('0')) + order.total_amount
                    
                    # Procesar items de la orden
                    for item in order.orderitem_set.all():
                        category = item.recipe.group.name if item.recipe and item.recipe.group else 'Sin categoría'
                        
                        # Stats por categoría
                        if category not in category_stats:
                            category_stats[category] = {'revenue': Decimal('0'), 'quantity': 0}
                        category_stats[category]['revenue'] += item.total_price
                        category_stats[category]['quantity'] += item.quantity
                        
                        # Stats por plato
                        dish_key = item.recipe.name if item.recipe else 'Sin receta'
                        if dish_key not in dish_stats:
                            dish_stats[dish_key] = {
                                'category': category,
                                'quantity': 0,
                                'revenue': Decimal('0'),
                                'unit_price': item.unit_price
                            }
                        dish_stats[dish_key]['quantity'] += item.quantity
                        dish_stats[dish_key]['revenue'] += item.total_price
                    
                    # Procesar pagos
                    for payment in order.payments.all():
                        method = payment.payment_method
                        payment_method_totals[method] = payment_method_totals.get(method, Decimal('0')) + payment.amount
                        
                except Exception as e:
                    print(f"⚠️ Error processing order {order.id}: {e}")
                    continue
            
            print("✅ Finished processing orders, calculating final stats...")
            
            # Calcular tiempo de servicio promedio
            average_service_time = sum(service_times) / len(service_times) if service_times else 0
            
            # Calcular porcentajes para categorías
            total_category_revenue = sum(cat['revenue'] for cat in category_stats.values())
            category_breakdown = []
            for category, stats in sorted(category_stats.items(), key=lambda x: x[1]['revenue'], reverse=True):
                percentage = (stats['revenue'] / total_category_revenue * 100) if total_category_revenue > 0 else 0
                category_breakdown.append({
                    'category': category,
                    'revenue': float(stats['revenue']),
                    'quantity': stats['quantity'],
                    'percentage': float(percentage)
                })
            
            # Top 10 platos
            top_dishes = []
            for dish, stats in sorted(dish_stats.items(), key=lambda x: x[1]['quantity'], reverse=True)[:10]:
                top_dishes.append({
                    'name': dish,
                    'category': stats['category'],
                    'quantity': stats['quantity'],
                    'revenue': float(stats['revenue']),
                    'unit_price': float(stats['unit_price'])
                })
            
            # Top 5 meseros
            waiter_performance = []
            for waiter, stats in sorted(waiter_revenue.items(), key=lambda x: x[1]['revenue'], reverse=True)[:5]:
                avg_ticket = stats['revenue'] / stats['orders'] if stats['orders'] > 0 else Decimal('0')
                waiter_performance.append({
                    'waiter': waiter,
                    'orders': stats['orders'],
                    'revenue': float(stats['revenue']),
                    'average_ticket': float(avg_ticket)
                })
            
            # Performance por zonas
            zone_performance = []
            for zone, stats in sorted(zone_revenue.items(), key=lambda x: x[1]['revenue'], reverse=True):
                tables_used = len(stats['tables'])
                avg_per_table = stats['revenue'] / tables_used if tables_used > 0 else Decimal('0')
                zone_performance.append({
                    'zone': zone,
                    'orders': stats['orders'],
                    'revenue': float(stats['revenue']),
                    'tables_used': tables_used,
                    'average_per_table': float(avg_per_table)
                })
            
            # Top 5 mesas
            top_tables = []
            for table, revenue in sorted(table_revenue.items(), key=lambda x: x[1], reverse=True)[:5]:
                top_tables.append({
                    'table': table,
                    'revenue': float(revenue)
                })
            
            # Distribución por método de pago
            total_payments = sum(payment_method_totals.values())
            payment_methods = []
            for method, amount in payment_method_totals.items():
                percentage = (amount / total_payments * 100) if total_payments > 0 else 0
                payment_methods.append({
                    'method': method,
                    'amount': float(amount),
                    'percentage': float(percentage)
                })
            
            print("✅ Dashboard data compiled successfully")
            
            # Respuesta consolidada
            response_data = {
                'date': selected_date.isoformat(),
                'summary': {
                    'total_orders': total_orders,
                    'total_revenue': float(total_revenue),
                    'average_ticket': float(average_ticket),
                    'average_service_time': float(average_service_time)
                },
                'category_breakdown': category_breakdown,
                'top_dishes': top_dishes,
                'waiter_performance': waiter_performance,
                'zone_performance': zone_performance,
                'top_tables': top_tables,
                'payment_methods': payment_methods
            }
            
            # Debug: imprimir resumen de datos
            print(f"✅ RESUMEN DE DATOS GENERADOS:")
            print(f"   📊 Órdenes: {total_orders}")
            print(f"   💰 Ingresos: {total_revenue}")
            print(f"   📈 Ticket promedio: {average_ticket}")
            print(f"   ⏱️ Tiempo servicio: {average_service_time:.1f} min")
            print(f"   🏷️ Categorías: {len(category_breakdown)}")
            print(f"   🍽️ Top platos: {len(top_dishes)}")
            print(f"   👤 Meseros: {len(waiter_performance)}")
            print(f"   🏢 Zonas: {len(zone_performance)}")
            print(f"   🪑 Mesas top: {len(top_tables)}")
            print(f"   💳 Métodos pago: {len(payment_methods)}")
            
            print(f"✅ Returning response with {total_orders} orders processed")
            return Response(response_data)
        
        except Exception as e:
            print(f"❌ CRITICAL ERROR in dashboard report: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({
                'error': f'Error processing dashboard request: {str(e)}',
                'date': None,
                'summary': {'total_orders': 0, 'total_revenue': 0, 'average_ticket': 0},
                'category_breakdown': [], 'top_dishes': [], 'waiter_performance': [],
                'zone_performance': [], 'top_tables': [], 'payment_methods': []
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    @action(detail=False, methods=['get'])
    def export_excel(self, request):
        """
        Exporta el reporte del dashboard a Excel (CSV como fallback)
        """
        try:
            print("🔍 Starting Excel/CSV export...")
            
            # Obtener datos usando el mismo método
            report_response = self.report(request)
            if report_response.status_code != 200:
                return report_response
            
            response_data = report_response.data
            print(f"✅ Got dashboard data for export: {response_data.get('date')}")
            
            # Intentar importar openpyxl
            try:
                import openpyxl
                from openpyxl.styles import Font, PatternFill, Alignment
                excel_available = True
                print("✅ openpyxl available - generating Excel file")
            except ImportError:
                excel_available = False
                print("⚠️ openpyxl not available - generating CSV file")
            
            if excel_available:
                # Generar Excel
                return self._generate_excel(response_data)
            else:
                # Generar CSV como fallback
                return self._generate_csv(response_data)
                
        except Exception as e:
            print(f"❌ Error generating export file: {str(e)}")
            import traceback
            traceback.print_exc()
            return Response({
                'error': f'Error generating export file: {str(e)}'
            }, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
    
    def _generate_csv(self, response_data):
        """
        Genera un CSV con los datos del dashboard
        """
        import csv
        from django.http import HttpResponse
        import io
        
        print("📊 Generating CSV export...")
        
        # Crear CSV en memoria
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Header del archivo
        writer.writerow([f"Dashboard de Ventas - {response_data['date']}"])
        writer.writerow([])
        
        # Resumen
        writer.writerow(['RESUMEN GENERAL'])
        writer.writerow(['Métrica', 'Valor'])
        writer.writerow(['Total de Órdenes', response_data['summary']['total_orders']])
        writer.writerow(['Ingresos Totales', f"S/ {response_data['summary']['total_revenue']:.2f}"])
        writer.writerow(['Ticket Promedio', f"S/ {response_data['summary']['average_ticket']:.2f}"])
        writer.writerow(['Tiempo Servicio Promedio', f"{response_data['summary']['average_service_time']:.1f} min"])
        writer.writerow([])
        
        # Categorías
        writer.writerow(['VENTAS POR CATEGORÍA'])
        writer.writerow(['Categoría', 'Ingresos', 'Cantidad', 'Porcentaje'])
        for cat in response_data['category_breakdown']:
            writer.writerow([
                cat['category'],
                f"S/ {cat['revenue']:.2f}",
                cat['quantity'],
                f"{cat['percentage']:.1f}%"
            ])
        writer.writerow([])
        
        # Top platos
        writer.writerow(['TOP PLATOS'])
        writer.writerow(['Ranking', 'Plato', 'Categoría', 'Cantidad', 'Ingresos', 'Precio Unitario'])
        for idx, dish in enumerate(response_data['top_dishes'], 1):
            writer.writerow([
                idx,
                dish['name'],
                dish['category'],
                dish['quantity'],
                f"S/ {dish['revenue']:.2f}",
                f"S/ {dish['unit_price']:.2f}"
            ])
        writer.writerow([])
        
        # Performance meseros
        writer.writerow(['PERFORMANCE MESEROS'])
        writer.writerow(['Mesero', 'Órdenes', 'Ingresos', 'Ticket Promedio'])
        for waiter in response_data['waiter_performance']:
            writer.writerow([
                waiter['waiter'],
                waiter['orders'],
                f"S/ {waiter['revenue']:.2f}",
                f"S/ {waiter['average_ticket']:.2f}"
            ])
        writer.writerow([])
        
        # Performance zonas
        writer.writerow(['PERFORMANCE ZONAS'])
        writer.writerow(['Zona', 'Órdenes', 'Ingresos', 'Mesas Usadas', 'Promedio por Mesa'])
        for zone in response_data['zone_performance']:
            writer.writerow([
                zone['zone'],
                zone['orders'],
                f"S/ {zone['revenue']:.2f}",
                zone['tables_used'],
                f"S/ {zone['average_per_table']:.2f}"
            ])
        writer.writerow([])
        
        # Métodos de pago
        writer.writerow(['MÉTODOS DE PAGO'])
        writer.writerow(['Método', 'Monto', 'Porcentaje'])
        for method in response_data['payment_methods']:
            writer.writerow([
                method['method'],
                f"S/ {method['amount']:.2f}",
                f"{method['percentage']:.1f}%"
            ])
        
        # Preparar respuesta
        response = HttpResponse(content_type='text/csv; charset=utf-8')
        filename = f"dashboard_ventas_{response_data['date']}.csv"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response.write(output.getvalue())
        
        print("✅ CSV file generated successfully")
        return response
    
    def _generate_excel(self, response_data):
        """
        Genera un archivo Excel con los datos del dashboard
        """
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from django.http import HttpResponse
        
        print("📊 Generating Excel export...")
        
        # Crear libro de Excel
        wb = openpyxl.Workbook()
        
        # Estilos
        header_font = Font(bold=True, color="FFFFFF", size=12)
        header_fill = PatternFill("solid", fgColor="366092")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Hoja 1: Resumen General
        ws_summary = wb.active
        ws_summary.title = "Resumen"
        
        # Título
        ws_summary.merge_cells('A1:D1')
        ws_summary['A1'] = f"Dashboard de Ventas - {response_data['date']}"
        ws_summary['A1'].font = Font(bold=True, size=16)
        ws_summary['A1'].alignment = Alignment(horizontal="center")
        
        # Resumen
        ws_summary['A3'] = "Métrica"
        ws_summary['B3'] = "Valor"
        ws_summary['A3'].font = header_font
        ws_summary['B3'].font = header_font
        ws_summary['A3'].fill = header_fill
        ws_summary['B3'].fill = header_fill
        
        summary_data = [
            ("Total de Órdenes", response_data['summary']['total_orders']),
            ("Ingresos Totales", f"S/ {response_data['summary']['total_revenue']:.2f}"),
            ("Ticket Promedio", f"S/ {response_data['summary']['average_ticket']:.2f}"),
            ("Tiempo Servicio Promedio", f"{response_data['summary']['average_service_time']:.1f} min")
        ]
        
        for idx, (metric, value) in enumerate(summary_data, start=4):
            ws_summary[f'A{idx}'] = metric
            ws_summary[f'B{idx}'] = value
        
        # Preparar respuesta HTTP
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        filename = f"dashboard_ventas_{response_data['date']}.xlsx"
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        
        # Guardar y retornar
        wb.save(response)
        print("✅ Excel file generated successfully")
        return response